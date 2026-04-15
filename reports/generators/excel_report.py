# -*- coding: utf-8 -*-
import logging
from datetime import datetime
from pathlib import Path

from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference

logger = logging.getLogger(__name__)

MEDIA_ROOT = Path(settings.MEDIA_ROOT)
REPORTS_DIR = MEDIA_ROOT / "reports"

COLOR_GREEN = "2E7D32"
COLOR_WHITE = "FFFFFF"
COLOR_LIGHT_GREEN = "E8F5E9"

def _apply_header_style(sheet):
    """Applies green background, white bold font to the first row."""
    header_font = Font(color=COLOR_WHITE, bold=True)
    header_fill = PatternFill(start_color=COLOR_GREEN, end_color=COLOR_GREEN, fill_type="solid")
    
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

def _apply_borders(sheet):
    """Applies borders to all cells with data."""
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = thin_border

def _autofit_columns(sheet):
    """Adjusts column widths based on content."""
    for column_cells in sheet.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = length + 2

def generate_detection_excel(detection_results) -> str:
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    filename = f"detection_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    output_path = REPORTS_DIR / filename

    wb = Workbook()
    
    # Sheet 1: Detection Results
    ws1 = wb.active
    ws1.title = "Tespit Sonuçları"
    ws1.append([
        "ID", "Meyve Türü", "Tespit Sayısı", "Ağaç Sayısı", "Ağaç Yaşı",
        "Tek Ağaç Ağırlığı (kg)", "Toplam Ağırlık (kg)", "Güven Skoru",
        "İşlem Süresi (sn)", "Model", "Eşik", "Tespit Tarihi"
    ])

    for detection in detection_results:
        ws1.append([
            detection.pk,
            detection.fruit_type,
            detection.detected_count,
            detection.tree_count,
            detection.tree_age,
            detection.weight,
            detection.total_weight,
            detection.confidence_score,
            detection.processing_time,
            detection.model_version,
            detection.threshold_used,
            detection.created_at.replace(tzinfo=None)  # Excel doesn't like timezone aware datetimes sometimes
        ])
    
    _apply_header_style(ws1)
    _apply_borders(ws1)
    _autofit_columns(ws1)
    ws1.freeze_panes = "A2"

    # Alternating row colors
    fill_light = PatternFill(start_color=COLOR_LIGHT_GREEN, end_color=COLOR_LIGHT_GREEN, fill_type="solid")
    for row in ws1.iter_rows(min_row=2):
        if row[0].row % 2 == 0:
            for cell in row:
                cell.fill = fill_light

    # Sheet 2: Summary
    ws2 = wb.create_sheet("Özet")
    total_records = len(detection_results)
    total_detections = sum(d.detected_count for d in detection_results)
    total_weight = sum(d.total_weight for d in detection_results)
    avg_confidence = sum(d.confidence_score for d in detection_results) / total_records if total_records else 0
    avg_processing_time = sum(d.processing_time for d in detection_results) / total_records if total_records else 0
    max_detections = max((d.detected_count for d in detection_results), default=0)
    min_detections = min((d.detected_count for d in detection_results), default=0)

    ws2.append(["İstatistik", "Değer"])
    ws2.append(["Toplam Kayıt", total_records])
    ws2.append(["Toplam Tespit", total_detections])
    ws2.append(["Toplam Ağırlık (kg)", total_weight])
    ws2.append(["Ortalama Güven Skoru", avg_confidence])
    ws2.append(["Ortalama İşlem Süresi (sn)", avg_processing_time])
    ws2.append(["Maksimum Tespit", max_detections])
    ws2.append(["Minimum Tespit", min_detections])

    _apply_header_style(ws2)
    _apply_borders(ws2)
    _autofit_columns(ws2)

    # Sheet 3: Group by Fruit Type
    ws3 = wb.create_sheet("Meyve Türüne Göre")
    ws3.append(["Meyve Türü", "Kayıt Sayısı", "Toplam Tespit", "Toplam Ağırlık (kg)"])

    fruit_stats = {}
    for d in detection_results:
        ft = d.fruit_type
        if ft not in fruit_stats:
            fruit_stats[ft] = {'count': 0, 'detections': 0, 'weight': 0}
        fruit_stats[ft]['count'] += 1
        fruit_stats[ft]['detections'] += d.detected_count
        fruit_stats[ft]['weight'] += d.total_weight

    row_idx = 2
    for fruit, stats in fruit_stats.items():
        ws3.append([fruit, stats['count'], stats['detections'], stats['weight']])
        row_idx += 1

    _apply_header_style(ws3)
    _apply_borders(ws3)
    _autofit_columns(ws3)

    # Bar Chart
    chart = BarChart()
    chart.title = "Meyve Türüne Göre Toplam Tespit"
    chart.x_axis.title = "Meyve Türü"
    chart.y_axis.title = "Tespit Sayısı"
    
    # Data is in column 3 (C), Categories in column 1 (A)
    data = Reference(ws3, min_col=3, min_row=1, max_row=len(fruit_stats)+1)
    cats = Reference(ws3, min_col=1, min_row=2, max_row=len(fruit_stats)+1)
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws3.add_chart(chart, "E2")

    wb.save(output_path)
    logger.info("Detection Excel generated: %s", output_path)
    return f"reports/{filename}"


def generate_drone_excel(project, analysis_data: dict) -> str:
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    filename = f"drone_report_{project.pk}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    output_path = REPORTS_DIR / filename

    wb = Workbook()

    # Sheet 1: Project Info
    ws1 = wb.active
    ws1.title = "Proje Bilgileri"
    ws1.append(["Alan", "Değer"])
    ws1.append(["Proje Adı", project.Title])
    ws1.append(["Çiftlik", project.Farm])
    ws1.append(["Tarla", project.Field])
    ws1.append(["Durum", project.State])
    ws1.append(["Algoritma", analysis_data.get('algorithm', '—')])
    ws1.append(["Analiz Tarihi", analysis_data.get('analysis_date', '—')])

    _apply_header_style(ws1)
    _apply_borders(ws1)
    _autofit_columns(ws1)

    # Sheet 2: Vegetation Stats
    ws2 = wb.create_sheet("İndeks İstatistikleri")
    ws2.append(["Metrik", "Değer"])
    for k, v in analysis_data.get('vegetation_stats', {}).items():
        ws2.append([k, v])
    
    _apply_header_style(ws2)
    _apply_borders(ws2)
    _autofit_columns(ws2)

    # Sheet 3: Stress Zones
    ws3 = wb.create_sheet("Stres Zonları")
    ws3.append(["Zone ID", "Stres Sınıfı", "Alan (ha)"])
    for zone in analysis_data.get('stress_zones', []):
        ws3.append([
            zone.get('zone_id'),
            zone.get('stress_class'),
            zone.get('area_ha')
        ])
    
    _apply_header_style(ws3)
    _apply_borders(ws3)
    _autofit_columns(ws3)

    # Sheet 4: Yield Prediction
    ws4 = wb.create_sheet("Verim Tahmini")
    ws4.append(["Metrik", "Tahmin"])
    for k, v in analysis_data.get('yield_prediction', {}).items():
        ws4.append([k, v])
        
    _apply_header_style(ws4)
    _apply_borders(ws4)
    _autofit_columns(ws4)

    # Sheet 5: Recommendations
    ws5 = wb.create_sheet("Öneriler")
    ws5.append(["Önem Derecesi", "Aksiyon"])
    for rec in analysis_data.get('recommendations', []):
        ws5.append([
            rec.get('severity'),
            rec.get('action')
        ])

    _apply_header_style(ws5)
    _apply_borders(ws5)
    _autofit_columns(ws5)

    wb.save(output_path)
    logger.info("Drone Excel generated: %s", output_path)
    return f"reports/{filename}"
